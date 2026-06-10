import base64
import io
import mimetypes
from typing import Any

from fastapi import HTTPException, UploadFile
from docx import Document as DocxDocument
from openpyxl import load_workbook

from ai_common import HAIKU_MODEL, call_claude, get_db, parse_json_response, to_json


SYSTEM_PROMPT = """
Eres un asistente experto en clasificar solicitudes operativas y administrativas dentro de una empresa.
Tu trabajo es elegir el workflow mas especifico y util para la solicitud del usuario.

Reglas:
- Debes elegir SOLO entre los workflows proporcionados.
- Prioriza el workflow mas especifico sobre uno generico.
- Usa el texto del usuario, la transcripcion de voz y la documentacion adjunta para inferir el workflow correcto.
- Si la documentacion contradice el texto breve del usuario, usa la interpretacion mas consistente con la evidencia total.
- Tambien debes extraer datos utiles para el formulario inicial del workflow elegido.
- Respeta los nombres exactos de los campos del formulario inicial al devolver prefillData.
- No inventes valores que no esten sustentados.
- Si un dato requerido no puede inferirse, agregalo en missingRequiredFields.
- Si no hay ningun workflow suficientemente adecuado, devuelve workflowId como null.

Responde SOLO JSON valido con esta forma:
{
  "workflowId": "string o null",
  "workflowName": "string o null",
  "confidence": 0.0,
  "reasoning": "explicacion corta",
  "detectedIntent": "resumen corto del pedido",
  "prefillData": {},
  "missingRequiredFields": ["campo1"],
  "suggestedQuestions": ["pregunta para completar faltantes"],
  "alternatives": [
    {"workflowId": "string", "workflowName": "string", "reason": "string corto"}
  ]
}
""".strip()


def process_workflow_router(prompt: str, company_id: str | None, files: list[UploadFile] | None = None) -> dict[str, Any]:
    request_text = (prompt or "").strip()
    if not request_text:
        raise HTTPException(status_code=400, detail="Debes enviar una descripcion de la solicitud")

    workflows = load_company_workflows(company_id)
    if not workflows:
        raise HTTPException(status_code=404, detail="No hay workflows disponibles para clasificar")

    content_blocks: list[dict[str, Any]] = [
        {
            "type": "text",
            "text": (
                f"Solicitud del usuario:\n{request_text}\n\n"
                f"Workflows disponibles:\n{to_json(workflows)}\n\n"
                "Analiza la solicitud y elige el mejor workflow posible."
            ),
        }
    ]

    for file in files or []:
        content_blocks.extend(_file_to_claude_blocks(file))

    raw = call_claude(
        system_prompt=SYSTEM_PROMPT,
        model=HAIKU_MODEL,
        max_tokens=1600,
        messages=[{"role": "user", "content": content_blocks}],
    )
    parsed = parse_json_response(raw)
    if not parsed:
        raise HTTPException(status_code=502, detail="La IA no devolvio una clasificacion valida")

    workflow_map = {workflow["id"]: workflow for workflow in workflows}
    workflow_id = parsed.get("workflowId")
    selected = workflow_map.get(workflow_id) if workflow_id else None
    if selected:
        parsed["workflowName"] = selected["name"]
        required_fields = {
            field["name"]
            for field in selected.get("entryFields", [])
            if field.get("required")
        }
        parsed["missingRequiredFields"] = [
            field_name
            for field_name in parsed.get("missingRequiredFields", [])
            if field_name in required_fields
        ]
        parsed["prefillData"] = {
            key: value
            for key, value in (parsed.get("prefillData") or {}).items()
            if key in {field["name"] for field in selected.get("entryFields", [])}
        }
    else:
        parsed["workflowId"] = None
        parsed["workflowName"] = None
        parsed["prefillData"] = {}
        parsed["missingRequiredFields"] = []

    if not isinstance(parsed.get("alternatives"), list):
        parsed["alternatives"] = []
    if not isinstance(parsed.get("suggestedQuestions"), list):
        parsed["suggestedQuestions"] = []
    return parsed


def load_company_workflows(company_id: str | None) -> list[dict[str, Any]]:
    db = get_db()
    workflow_query: dict[str, Any] = {}
    if company_id:
        workflow_query["companyId"] = company_id
    workflows = list(db["workflow"].find(workflow_query, {"name": 1, "description": 1, "companyId": 1}))
    if not workflows:
        return []

    workflow_ids = [str(item["_id"]) for item in workflows]
    nodos = list(db["workflow_nodo"].find({"workflowId": {"$in": workflow_ids}}))
    transitions = list(db["workflow_transition"].find({"workflowId": {"$in": workflow_ids}}))
    forms = list(db["form_definitions"].find({"nodoId": {"$in": [str(nodo["_id"]) for nodo in nodos]}}))

    nodos_by_workflow: dict[str, list[dict[str, Any]]] = {}
    for nodo in nodos:
        nodos_by_workflow.setdefault(str(nodo.get("workflowId")), []).append(nodo)

    transitions_by_workflow: dict[str, list[dict[str, Any]]] = {}
    for transition in transitions:
        transitions_by_workflow.setdefault(str(transition.get("workflowId")), []).append(transition)

    form_by_nodo = {str(form.get("nodoId")): form for form in forms}

    result: list[dict[str, Any]] = []
    for workflow in workflows:
        workflow_id = str(workflow["_id"])
        workflow_nodos = sorted(
            nodos_by_workflow.get(workflow_id, []),
            key=lambda item: int(item.get("order", 0)),
        )
        workflow_transitions = transitions_by_workflow.get(workflow_id, [])
        entry_nodo = _resolve_entry_nodo(workflow_nodos, workflow_transitions)
        entry_form = form_by_nodo.get(str(entry_nodo.get("_id"))) if entry_nodo else None
        result.append({
            "id": workflow_id,
            "name": workflow.get("name"),
            "description": workflow.get("description") or "",
            "entryNodoName": entry_nodo.get("name") if entry_nodo else None,
            "entryRequiresForm": bool(entry_nodo and entry_nodo.get("requiresForm")),
            "entryFields": _map_entry_fields(entry_form),
            "nodoSummary": [
                {
                    "name": nodo.get("name"),
                    "description": nodo.get("description") or "",
                    "type": nodo.get("nodeType") or "proceso",
                }
                for nodo in workflow_nodos[:12]
            ],
        })
    return result


def _resolve_entry_nodo(nodos: list[dict[str, Any]], transitions: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not nodos:
        return None
    nodo_by_id = {str(nodo["_id"]): nodo for nodo in nodos}
    start = next((nodo for nodo in nodos if str(nodo.get("nodeType", "")).lower() == "inicio"), None)
    if start:
        first_transition = next(
            (transition for transition in transitions if str(transition.get("fromNodoId")) == str(start["_id"])),
            None,
        )
        if first_transition:
            return nodo_by_id.get(str(first_transition.get("toNodoId")))
    return next((nodo for nodo in nodos if str(nodo.get("nodeType", "")).lower() != "inicio"), nodos[0])


def _map_entry_fields(form_definition: dict[str, Any] | None) -> list[dict[str, Any]]:
    if not form_definition:
        return []
    fields = []
    for raw_field in form_definition.get("fields", []) or []:
        field_type = str(raw_field.get("type") or "TEXT").upper()
        fields.append({
            "name": raw_field.get("name"),
            "type": field_type,
            "required": bool(raw_field.get("isRequired") or raw_field.get("required")),
            "columns": [
                {
                    "name": column.get("name"),
                    "type": str(column.get("type") or "TEXT").upper(),
                }
                for column in raw_field.get("columns", []) or []
            ] if field_type == "GRID" else [],
        })
    return fields


def _file_to_claude_blocks(file: UploadFile) -> list[dict[str, Any]]:
    filename = file.filename or "archivo"
    content = file.file.read()
    if not content:
        return []

    media_type = file.content_type or mimetypes.guess_type(filename)[0] or "application/octet-stream"
    lower_name = filename.lower()

    if media_type == "application/pdf" or lower_name.endswith(".pdf"):
        return [
            {
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": "application/pdf",
                    "data": base64.b64encode(content).decode("utf-8"),
                },
            },
            {"type": "text", "text": f"Documento adjunto: {filename}"},
        ]

    if media_type in {"image/jpeg", "image/png", "image/gif", "image/webp"}:
        return [
            {
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": media_type,
                    "data": base64.b64encode(content).decode("utf-8"),
                },
            },
            {"type": "text", "text": f"Imagen adjunta: {filename}"},
        ]

    extracted_text = _extract_text_from_file(filename, content, media_type)
    if extracted_text:
        return [{"type": "text", "text": f"Contenido del archivo {filename}:\n{extracted_text[:15000]}"}]

    return [{"type": "text", "text": f"Se adjunto el archivo {filename}, pero no se pudo extraer texto legible de su formato."}]


def _extract_text_from_file(filename: str, content: bytes, media_type: str) -> str:
    lower_name = filename.lower()
    try:
        if lower_name.endswith((".txt", ".md", ".csv")) or media_type.startswith("text/"):
            return content.decode("utf-8", errors="ignore")

        if lower_name.endswith(".docx") or media_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            document = DocxDocument(io.BytesIO(content))
            return "\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text.strip())

        if lower_name.endswith(".xlsx") or media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            rows: list[str] = []
            for sheet in workbook.worksheets:
                rows.append(f"[Hoja: {sheet.title}]")
                for row in sheet.iter_rows(values_only=True):
                    values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                    if values:
                        rows.append(" | ".join(values))
            return "\n".join(rows)
    except Exception:
        return ""
    return ""
